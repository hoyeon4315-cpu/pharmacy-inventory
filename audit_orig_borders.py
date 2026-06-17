# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter

ORIG = Path(r"C:\Users\duih\Desktop\월별 보고자료 (to 파트장님)\항암제 TPN 조제통계(20년  월)-보고용.xlsx")
wb = openpyxl.load_workbook(ORIG)
ws = wb['항암제']

def sides(addr):
    c = ws[addr]
    b = c.border
    out = {}
    for s in ['left', 'right', 'top', 'bottom']:
        x = getattr(b, s)
        out[s] = x.style if x and x.style else '-'
    return out

print('=== row2 B-N ===')
for col in range(2, 15):
    a = f'{get_column_letter(col)}2'
    print(a, sides(a))

print('=== row3 sample ===')
for col in range(1, 15):
    a = f'{get_column_letter(col)}3'
    print(a, sides(a))

print('=== row10 ===')
for col in range(1, 15):
    a = f'{get_column_letter(col)}10'
    print(a, sides(a))

print('=== row61 TOTAL ===')
for col in range(1, 15):
    a = f'{get_column_letter(col)}61'
    print(a, sides(a))

# count cells with missing thin in grid 3-60, B-N
missing = []
for r in range(3, 61):
    for col in range(2, 14):
        a = f'{get_column_letter(col)}{r}'
        s = sides(a)
        for k, v in s.items():
            if v == '-':
                missing.append(f'{a}.{k}')
print('missing border sides in B3:M60:', len(missing), missing[:20])
wb.close()