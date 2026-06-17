# -*- coding: utf-8 -*-
"""원본(20년 월) 기준 — 표 전체 얇은 실선 격자 + 바깥만 굵게."""
from copy import copy
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

ORIG = Path(r"C:\Users\duih\Desktop\월별 보고자료 (to 파트장님)\항암제 TPN 조제통계(20년  월)-보고용.xlsx")
OUT_PUBLIC = Path(__file__).parent / "public" / "stats-report-template.xlsx"
OUT_LOCAL = Path(__file__).parent / "stats-report-template.xlsx"

DEPTS = ['GS', 'GIC', 'IMH', 'IMR', 'OG', 'IMJ', 'UR', 'NR', 'OR', 'IMN', '기타']
DRUGS = [
    '5-FU', 'Abraxane', 'Alimta', 'Babencio', 'Bencord', 'Bleomycin', 'Caelyx', 'Campto', 'Camtobell',
    'Carboplatin', 'Cisplan', 'Cyramza', 'Cytarabine', 'Dacogen', 'Dazalex', 'Ditaxel', 'Doxorubin',
    'Endoxan', 'Enhertu', 'EPS', 'Erbitux', 'Gazyva', 'Gemzar', 'Halaven', 'Herceptin SC', 'Herzuma',
    'Holoxan', 'Imfinzi', 'Jemperli', 'Keytruda', 'Kyprolis', 'Mabthera', 'Mabthera SC',
    'Methotrexate(5g)', 'Mitomycin-C', 'Mitron', 'MTX', 'Navelbine', 'Onbevzy', 'Onivyde', 'Opdivo',
    'Oxalitin', 'Padexol', 'Perjeta', 'Samfenet', 'Sylvant', 'Taxol 30mg', 'Tecentriq', 'Tevimbra',
    'Topocan', 'Trisenox', 'Trodelvy', 'Velbastine', 'Velcade', 'Vincristine', 'Vyloy', 'Yervoy',
    'Zaltrap', 'Zavedos', 'Zepzelca',
]

BEIGE = 'FFEEECE1'
GRAY = 'FFD9D9D9'
THIN = Side(style='thin', color='000000')
MED = Side(style='medium', color='000000')


def grid_border(r, c, hdr_row, total_row, tpn_row, sum_col=13):
    """표 안: 전부 thin. 바깥 테두리만 medium (원본과 동일한 느낌)."""
    t = b = l = ri = THIN
    if c == 1:
        l = ri = MED
    if c == sum_col:
        l = ri = MED
    if r == hdr_row:
        t = MED
    if r == total_row:
        t = MED
    if r == tpn_row:
        b = MED
    return Border(left=l, top=t, right=ri, bottom=b)


def build():
    ref_wb = openpyxl.load_workbook(ORIG)
    ref = ref_wb['항암제']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '항암제'

    for col in range(1, 14):
        letter = get_column_letter(col)
        if letter in ref.column_dimensions:
            ws.column_dimensions[letter].width = ref.column_dimensions[letter].width
    ws.row_dimensions[1].height = ref.row_dimensions[1].height
    ws.row_dimensions[2].height = ref.row_dimensions[2].height

    hdr_row = 2
    first_data = 3
    last_data = first_data + len(DRUGS) - 1
    total_row = last_data + 1
    tpn_row = total_row + 1
    footer_row = tpn_row + 2
    sum_col = 13
    dept_last = 12

    align = Alignment(horizontal='center', vertical='center')
    font_title = copy(ref['A1'].font)
    font_hdr = Font(name='맑은 고딕', sz=11)
    font_drug = Font(name='맑은 고딕', sz=10)
    font_data = Font(name='맑은 고딕', sz=11)
    beige = PatternFill(patternType='solid', fgColor=BEIGE)
    gray = PatternFill(patternType='solid', fgColor=GRAY)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    t = ws['A1']
    t.value = 'YYYY년 M월 항암제 조제통계'
    t.font = font_title
    t.alignment = align
    t.border = Border(left=MED, top=MED, right=MED, bottom=THIN)

    for c in range(1, sum_col + 1):
        cell = ws.cell(hdr_row, c)
        cell.alignment = align
        cell.font = font_hdr
        cell.fill = beige
        cell.border = grid_border(hdr_row, c, hdr_row, total_row, tpn_row, sum_col)
    ws['A2'].value = None
    for i, d in enumerate(DEPTS, start=2):
        ws.cell(hdr_row, i, d)
    ws.cell(hdr_row, sum_col, '합계')

    for idx, drug in enumerate(DRUGS):
        r = first_data + idx
        a = ws.cell(r, 1, drug)
        a.font = font_drug
        a.alignment = align
        a.border = grid_border(r, 1, hdr_row, total_row, tpn_row, sum_col)
        for c in range(2, dept_last + 1):
            cell = ws.cell(r, c)
            cell.font = font_data
            cell.alignment = align
            cell.border = grid_border(r, c, hdr_row, total_row, tpn_row, sum_col)
            cell.value = None
        sc = ws.cell(r, sum_col)
        sc.font = font_data
        sc.alignment = align
        sc.fill = beige
        sc.border = grid_border(r, sum_col, hdr_row, total_row, tpn_row, sum_col)
        sc.value = f'=SUM(B{r}:{get_column_letter(dept_last)}{r})'

    for c in range(1, sum_col + 1):
        cell = ws.cell(total_row, c)
        cell.alignment = align
        cell.font = font_data
        cell.border = grid_border(total_row, c, hdr_row, total_row, tpn_row, sum_col)
        if c == 1:
            cell.fill = gray
            cell.value = 'TOTAL'
        elif c == sum_col:
            cell.fill = beige
            cell.value = f'=SUM(B{total_row}:{get_column_letter(dept_last)}{total_row})'
        else:
            cell.fill = gray
            letter = get_column_letter(c)
            cell.value = f'=SUM({letter}{first_data}:{letter}{last_data})'

    for c in range(1, sum_col + 1):
        cell = ws.cell(tpn_row, c)
        cell.alignment = align
        cell.font = font_data
        cell.border = grid_border(tpn_row, c, hdr_row, total_row, tpn_row, sum_col)
        if c == 1:
            cell.value = '특수TPN'
        elif c == 2:
            cell.value = '0건'
        elif c == sum_col:
            cell.fill = beige
            cell.value = f'=SUM(B{tpn_row}:{get_column_letter(dept_last)}{tpn_row})'
        else:
            cell.value = None

    fa = ws.cell(footer_row, 1, '외래주사실 :')
    fa.font = font_data
    fa.alignment = align
    ws.cell(footer_row, 2, '')

    ws.freeze_panes = 'A3'
    ref_wb.close()
    wb.save(OUT_PUBLIC)
    wb.save(OUT_LOCAL)
    print('template grid OK', OUT_PUBLIC)


if __name__ == '__main__':
    build()