# -*- coding: utf-8 -*-
"""Compare original template colors vs current deploy template."""
import zipfile
import re
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

ORIG = Path(r"C:\Users\duih\Desktop\월별 보고자료 (to 파트장님)\항암제 TPN 조제통계(20년  월)-보고용.xlsx")
CUR = Path(r"C:\Users\duih\Desktop\이호연\재고관리 프로그렘\public\stats-report-template.xlsx")

THEME_NAMES = ["dk1","lt1","dk2","lt2","accent1","accent2","accent3","accent4","accent5","accent6","hlink","folHlink"]

def theme_colors(path):
    with zipfile.ZipFile(path) as z:
        xml = z.read("xl/theme/theme1.xml").decode("utf-8")
    # parse clrScheme order
    order = ["dk1","lt1","dk2","lt2","accent1","accent2","accent3","accent4","accent5","accent6","hlink","folHlink"]
    colors = {}
    for name in order:
        m = re.search(rf"<a:{name}><a:srgbClr val=\"([0-9A-Fa-f]{{6}})\"", xml)
        if m:
            colors[name] = "#" + m.group(1).upper()
        else:
            m2 = re.search(rf"<a:{name}><a:sysClr val=\"(\w+)\" lastClr=\"([0-9A-Fa-f]{{6}})\"", xml)
            if m2:
                colors[name] = "#" + m2.group(2).upper() + f" (sys {m2.group(1)})"
    return colors

def resolve_theme(theme_idx, tint, palette):
    # OOXML theme index mapping used by Excel fgColor @theme
    key_map = {0:"lt1", 1:"dk1", 2:"dk2", 3:"lt2", 4:"accent1", 5:"accent2", 6:"accent3", 7:"accent4", 8:"accent5", 9:"accent6"}
    key = key_map.get(theme_idx, f"?{theme_idx}")
    base = palette.get(key, "?")
    return f"theme{theme_idx}({key}) tint={tint} -> {base}"

def cell_style(path, addrs):
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    palette = theme_colors(path)
    print(f"\n=== {path.name} === sheet={ws.title} freeze={ws.freeze_panes}")
    print("Theme palette:", palette)
    for addr in addrs:
        c = ws[addr]
        fg = c.fill.fgColor if c.fill else None
        if fg and fg.type == "theme":
            desc = resolve_theme(fg.theme, fg.tint, palette)
        elif fg and fg.rgb:
            desc = f"rgb {fg.rgb}"
        elif fg and fg.indexed is not None:
            desc = f"indexed {fg.indexed}"
        else:
            desc = "no fill"
        font = c.font
        fcol = font.color.rgb if font.color and font.color.rgb else (f"theme{font.color.theme}" if font.color and font.color.theme is not None else None)
        print(f"  {addr:4} val={repr(c.value)[:28]:28} fill={desc:45} fontColor={fcol}")
    wb.close()

def list_fills(path):
    with zipfile.ZipFile(path) as z:
        styles = z.read("xl/styles.xml").decode("utf-8")
    fills = re.findall(r"<fill>(.*?)</fill>", styles, re.S)
    print(f"\n{path.name} fill definitions ({len(fills)}):")
    for i, f in enumerate(fills):
        if "patternFill" in f and "none" not in f:
            print(f"  [{i}] {f.strip()}")

addrs = ["A1","B2","C2","L2","M2","A3","B3","K3","L3","M3","A63","B63","K63","L63","M63","A64","B64"]

for p in [ORIG, CUR]:
    if p.exists():
        cell_style(p, addrs)
        list_fills(p)
    else:
        print("MISSING", p)