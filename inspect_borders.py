# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path

ORIG = Path(r"C:\Users\duih\Desktop\월별 보고자료 (to 파트장님)\항암제 TPN 조제통계(20년  월)-보고용.xlsx")
CUR = Path(r"C:\Users\duih\Desktop\이호연\재고관리 프로그렘\public\stats-report-template.xlsx")

def border_str(b):
    if not b:
        return 'none'
    parts = []
    for side in ['left', 'right', 'top', 'bottom']:
        s = getattr(b, side)
        if s and s.style:
            parts.append(f"{side[0].upper()}:{s.style}")
    return ' '.join(parts) if parts else 'none'

for path in [ORIG, CUR]:
    wb = openpyxl.load_workbook(path)
    ws = wb['항암제']
    print(f"\n=== {path.name} ===")
    for addr in ['A1','B2','C2','L2','M2','A3','B3','C3','L3','M3','A10','M10','A62','B62','M62','A63','B63','M63','A64']:
        if ws[addr].value is None and addr not in ['B3','C3','L3','M3','A3']:
            pass
        c = ws[addr]
        print(f"  {addr:4} val={str(c.value)[:12]:12} border={border_str(c.border)}")
    wb.close()