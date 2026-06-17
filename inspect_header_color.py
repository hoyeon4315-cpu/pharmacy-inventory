# -*- coding: utf-8 -*-
import zipfile
import re
from pathlib import Path
import openpyxl
from openpyxl.styles.colors import COLOR_INDEX

ORIG = Path(r"C:\Users\duih\Desktop\월별 보고자료 (to 파트장님)\항암제 TPN 조제통계(20년  월)-보고용.xlsx")

with zipfile.ZipFile(ORIG) as z:
    styles_xml = z.read("xl/styles.xml").decode("utf-8")
    sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")

fills = re.findall(r"<fill>(.*?)</fill>", styles_xml, re.S)
cell_xfs = re.findall(r"<xf ([^/>]*)/>", styles_xml)

def get_xf(sid):
    return cell_xfs[int(sid)]

for cell in ["A2", "B2", "C2", "M2", "N2", "N3", "B3"]:
    m = re.search(r'r="%s"[^>]*s="(\d+)"' % cell, sheet_xml)
    if not m:
        print(cell, "no style id")
        continue
    sid = int(m.group(1))
    xf = get_xf(sid)
    print(f"\n{cell} xf[{sid}] = {xf}")
    fill_m = re.search(r'fillId="(\d+)"', xf)
    if fill_m:
        fid = int(fill_m.group(1))
        print(f"  fill[{fid}] = {fills[fid].strip()}")

wb = openpyxl.load_workbook(ORIG)
ws = wb["항암제"]
for addr in ["B2", "N2", "N3", "B3"]:
    c = ws[addr]
    fg = c.fill.fgColor
    print(f"\nopenpyxl {addr}: pattern={c.fill.patternType} type={fg.type} theme={fg.theme} tint={fg.tint} indexed={fg.indexed} rgb={fg.rgb}")
    if fg.indexed is not None and str(fg.indexed).isdigit():
        idx = int(fg.indexed)
        if idx < len(COLOR_INDEX):
            print(f"  indexed resolved: {COLOR_INDEX[idx]}")

# compare recent monthly files
for p in [
    Path(r"C:\Users\duih\Desktop\월별 보고자료 (to 파트장님)\2025\202505\파트장\항암제 TPN 조제통계(25년 05월)-보고용.xlsx"),
    Path(r"C:\Users\duih\Desktop\이호연\월마감\202604\파트장\항암제 TPN 조제통계(26년 4월)-보고용.xlsx"),
]:
    if p.exists():
        w = openpyxl.load_workbook(p)
        s = w["항암제"]
        fg = s["B2"].fill.fgColor
        print(f"\n{p.name} B2: theme={fg.theme} tint={fg.tint} rgb={fg.rgb}")
        w.close()