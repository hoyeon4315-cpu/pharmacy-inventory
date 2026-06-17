# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path

ORIG = Path(r"C:\Users\duih\Desktop\월별 보고자료 (to 파트장님)\항암제 TPN 조제통계(20년  월)-보고용.xlsx")
wb = openpyxl.load_workbook(ORIG)
ws = wb["항암제"]
print("dims", ws.dimensions, "max_row", ws.max_row, "max_col", ws.max_column)
print("freeze", ws.freeze_panes, "autofilter", ws.auto_filter.ref if ws.auto_filter else None)

# row2 headers
headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
print("row2:", headers)

# drugs col A
drugs = []
for r in range(3, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v:
        drugs.append((r, v))
print("drug rows", len(drugs), "first5", drugs[:5], "last5", drugs[-5:])

# cells with colored fill in first 70 rows
for r in range(1, min(70, ws.max_row + 1)):
    for c in range(1, 14):
        cell = ws.cell(r, c)
        fg = cell.fill.fgColor
        if cell.fill.patternType and cell.fill.patternType != 'none' and fg and fg.type == 'theme':
            print(f"  colored {cell.coordinate} theme={fg.theme} tint={fg.tint} val={cell.value}")

# sample M column
for r in [2,3,4,10,62,63,64,65,66]:
    if r <= ws.max_row:
        print(f"row{r} A={ws.cell(r,1).value} M={ws.cell(r,13).value}")
wb.close()